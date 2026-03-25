#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Standalone CombineProcess test runner.

Default behavior is dry-run. Use --save to persist evidence and update
the study-specific Markdown/XLSX test documents.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from contextlib import contextmanager, redirect_stdout
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl import load_workbook


def find_repo_root(start: Path) -> Path:
    for candidate in (start, *start.parents):
        if (candidate / "VC_BC01_constant.py").exists():
            return candidate
    raise FileNotFoundError("Could not locate repository root.")


CURRENT_DIR = Path(__file__).resolve().parent
REPO_ROOT = find_repo_root(CURRENT_DIR)
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import VC_BC04_operateType as operate_type

# Configure your study ID here
STUDY_ID = "example_study"
study_module = f"studySpecific.{STUDY_ID}.VC_BC05_studyFunctions"
import importlib
study_functions = importlib.import_module(study_module)


STUDY_DIR = REPO_ROOT / "studySpecific" / STUDY_ID
EVIDENCE_ROOT = STUDY_DIR / "testEvidence" / "CombineProcess"
DD_MD_PATH = STUDY_DIR / "detailed_design_CombineProcess.md"
DD_XLSX_PATH = STUDY_DIR / "detailed_design_CombineProcess.xlsx"
TE_MD_PATH = STUDY_DIR / "test_record_CombineProcess.md"
TE_XLSX_PATH = STUDY_DIR / "test_record_CombineProcess.xlsx"
TM_MD_PATH = STUDY_DIR / "traceability_matrix_CombineProcess.md"
TM_XLSX_PATH = STUDY_DIR / "traceability_matrix_CombineProcess.xlsx"
RD_MD_PATH = STUDY_DIR / "requirements_CombineProcess.md"
BD_MD_PATH = STUDY_DIR / "basic_design_CombineProcess.md"

PASS = "Pass"
FAIL = "Fail"
TM_STATUS = "In Progress"
ALLOWED_DM_COLUMNS = {
    "SUBJID",
    "SEXCD",
    "AGE",
    "LSVDAT",
    "DTHDAT",
    "RFENDAT",
    "DTHFLG",
}


@dataclass
class TestResult:
    tc_id: str
    title: str
    requirement_ids: list[str]
    expected: str
    status: str
    actual: str
    remark: str
    evidence_ids: list[str]
    evidence_docs: dict[str, str] = field(default_factory=dict)
    evidence_paths: dict[str, str] = field(default_factory=dict)

    def status_for_docs(self) -> str:
        return PASS if self.status == PASS else FAIL

    def evidence_cell_markdown(self) -> str:
        return "<br>".join(self.evidence_paths[evidence_id] for evidence_id in self.evidence_ids)

    def evidence_cell_workbook(self) -> str:
        return "\n".join(self.evidence_paths[evidence_id] for evidence_id in self.evidence_ids)


@dataclass
class RunSummary:
    generated_at: str
    mode: str
    execution_date: str
    environment: str
    evidence_directory: str | None
    pass_count: int
    fail_count: int
    overall_status: str
    results: list[dict[str, Any]]
    post_save_validation: dict[str, Any] = field(default_factory=dict)


class TestContext:
    def __init__(self) -> None:
        self._actual_dm: pd.DataFrame | None = None
        self._actual_tme_registration: pd.DataFrame | None = None

    def actual_dm(self) -> pd.DataFrame:
        if self._actual_dm is None:
            self._actual_dm = study_functions.DM()
        return self._actual_dm.copy()

    def actual_tme_registration(self) -> pd.DataFrame:
        if self._actual_tme_registration is None:
            self._actual_tme_registration = study_functions.filter_df_by_field(
                "TME", EventId="AT REGISTRATION"
            )
        return self._actual_tme_registration.copy()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run CombineProcess tests.")
    parser.add_argument(
        "--save",
        action="store_true",
        help="Persist evidence and update DD/TE/TM Markdown and XLSX files.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Explicitly run without saving artifacts.",
    )
    return parser.parse_args()


def now_local() -> datetime:
    return datetime.now().astimezone()


def format_date_jp(value: date) -> str:
    return value.strftime("%Y年%m月%d日")


def format_date_iso(value: date) -> str:
    return value.isoformat()


def sanitize_cell_text(value: str) -> str:
    return value.replace("|", "/").replace("\r\n", "\n").replace("\n", "<br>")


def normalize_doc_text(value: str) -> str:
    return value.replace("<br>", "\n").replace("\r\n", "\n").strip()


def csv_block(df: pd.DataFrame) -> str:
    return df.to_csv(index=False).strip()


def markdown_table_row(values: list[str]) -> str:
    sanitized = [sanitize_cell_text(value) for value in values]
    return "| " + " | ".join(sanitized) + " |"


def split_markdown_row(line: str) -> list[str]:
    return [part.strip() for part in line.strip().strip("|").split("|")]


def replace_markdown_row(line: str, updates: dict[int, str]) -> str:
    cells = split_markdown_row(line)
    for index, value in updates.items():
        if index < len(cells):
            cells[index] = value
    return markdown_table_row(cells)


def update_markdown_file(path: Path, updater: Any) -> None:
    original_text = path.read_text(encoding="utf-8")
    newline = "\r\n" if "\r\n" in original_text else "\n"
    updated_text = updater(original_text)
    path.write_text(updated_text.replace("\n", newline), encoding="utf-8")


def environment_string() -> str:
    return (
        f"Python {sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro} / "
        f"pandas {pd.__version__} / openpyxl {openpyxl.__version__}"
    )


def evidence_relpath(path: Path) -> str:
    return path.relative_to(REPO_ROOT).as_posix()


def build_evidence_markdown(
    result: TestResult,
    execution_date: str,
    mode: str,
    extra_sections: list[tuple[str, str]] | None = None,
) -> str:
    mode_display = {"save": "保存実行", "dry-run": "ドライラン"}.get(mode, mode)
    lines = [
        f"# {result.evidence_ids[0]}",
        "",
        f"- テストID: `{result.tc_id}`",
        f"- テストケース: {result.title}",
        f"- 対応要件: {', '.join(result.requirement_ids)}",
        f"- 実施日: {execution_date}",
        f"- 実行モード: {mode_display}",
        f"- 判定: **{result.status}**",
        "",
        "## 期待結果",
        result.expected,
        "",
        "## 実測結果",
        result.actual,
        "",
    ]
    if result.remark:
        lines.extend(["## 備考", result.remark, ""])
    if extra_sections:
        for title, content in extra_sections:
            lines.extend([f"## {title}", content, ""])
    return "\n".join(lines).rstrip() + "\n"


@contextmanager
def patched_format_dataset(dataset_map: dict[str, pd.DataFrame]):
    original = study_functions.getFormatDataset
    study_functions.getFormatDataset = lambda *args, **kwargs: dataset_map
    try:
        yield
    finally:
        study_functions.getFormatDataset = original


def find_row_by_first_value(worksheet: Any, value: str, first_col: int = 1) -> int:
    for row_idx in range(1, worksheet.max_row + 1):
        if worksheet.cell(row_idx, first_col).value == value:
            return row_idx
    raise KeyError(f"Row with value '{value}' was not found in sheet '{worksheet.title}'.")


def find_label_row(worksheet: Any, label: str, label_col: int) -> int:
    for row_idx in range(1, worksheet.max_row + 1):
        if worksheet.cell(row_idx, label_col).value == label:
            return row_idx
    raise KeyError(f"Label '{label}' was not found in sheet '{worksheet.title}'.")


def io_string() -> Any:
    from io import StringIO

    return StringIO()


def update_dd_markdown(results: dict[str, TestResult], updated_date: date) -> None:
    def updater(text: str) -> str:
        updated_lines = []
        for line in text.splitlines():
            stripped = line.strip()
            if stripped.startswith("| 最終更新日 |"):
                line = markdown_table_row(["最終更新日", format_date_jp(updated_date)])
            elif stripped.startswith("| TC-"):
                cells = split_markdown_row(line)
                tc_id = cells[0]
                if tc_id in results:
                    line = replace_markdown_row(line, {6: results[tc_id].status_for_docs()})
            updated_lines.append(line)
        return "\n".join(updated_lines) + "\n"

    update_markdown_file(DD_MD_PATH, updater)


def update_dd_workbook(results: dict[str, TestResult], updated_date: date) -> None:
    workbook = load_workbook(DD_XLSX_PATH)
    cover_sheet = workbook["表紙"]
    cover_row = find_label_row(cover_sheet, "最終更新日", 2)
    cover_sheet.cell(cover_row, 3).value = updated_date

    sheet = workbook["4.テスト仕様"]
    for tc_id, result in results.items():
        row_idx = find_row_by_first_value(sheet, tc_id, 1)
        sheet.cell(row_idx, 7).value = result.status_for_docs()

    workbook.save(DD_XLSX_PATH)


def update_te_markdown(
    results: dict[str, TestResult],
    updated_date: date,
    environment: str,
    summary_text: str,
    defect_row: list[str],
) -> None:
    def updater(text: str) -> str:
        updated_lines = []
        for line in text.splitlines():
            stripped = line.strip()
            if stripped.startswith("| 最終更新日 |"):
                line = markdown_table_row(["最終更新日", format_date_jp(updated_date)])
            elif stripped.startswith("| 実施日 |"):
                line = markdown_table_row(["実施日", format_date_iso(updated_date)])
            elif stripped.startswith("| 実施環境 |"):
                line = markdown_table_row(["実施環境", environment])
            elif stripped.startswith("| 実施結果サマリ |"):
                line = markdown_table_row(["実施結果サマリ", summary_text])
            elif stripped.startswith("| TC-"):
                cells = split_markdown_row(line)
                tc_id = cells[0]
                if tc_id in results:
                    result = results[tc_id]
                    line = markdown_table_row(
                        [
                            tc_id,
                            cells[1],
                            cells[2],
                            result.actual,
                            result.status_for_docs(),
                            cells[5],
                            result.evidence_cell_markdown(),
                            result.remark,
                        ]
                    )
            elif stripped.startswith("| DEF-001 |"):
                line = markdown_table_row(defect_row)
            updated_lines.append(line)
        return "\n".join(updated_lines) + "\n"

    update_markdown_file(TE_MD_PATH, updater)


def update_te_workbook(
    results: dict[str, TestResult],
    updated_date: date,
    environment: str,
    summary_text: str,
    defect_row: list[str],
) -> None:
    workbook = load_workbook(TE_XLSX_PATH)

    cover_sheet = workbook["表紙"]
    cover_row = find_label_row(cover_sheet, "最終更新日", 2)
    cover_sheet.cell(cover_row, 3).value = updated_date

    summary_sheet = workbook["1.実施概要"]
    summary_sheet.cell(find_label_row(summary_sheet, "実施日", 1), 2).value = format_date_iso(updated_date)
    summary_sheet.cell(find_label_row(summary_sheet, "実施環境", 1), 2).value = environment
    summary_sheet.cell(find_label_row(summary_sheet, "実施結果サマリ", 1), 2).value = summary_text

    result_sheet = workbook["2.テスト結果詳細"]
    for tc_id, result in results.items():
        row_idx = find_row_by_first_value(result_sheet, tc_id, 1)
        result_sheet.cell(row_idx, 4).value = result.actual
        result_sheet.cell(row_idx, 5).value = result.status_for_docs()
        result_sheet.cell(row_idx, 7).value = result.evidence_cell_workbook()
        result_sheet.cell(row_idx, 8).value = result.remark

    defect_sheet = workbook["3.不具合・課題記録"]
    defect_row_idx = find_row_by_first_value(defect_sheet, "DEF-001", 1)
    for col_idx, value in enumerate(defect_row[1:], start=2):
        defect_sheet.cell(defect_row_idx, col_idx).value = value

    workbook.save(TE_XLSX_PATH)


def update_tm_markdown(updated_date: date) -> None:
    def updater(text: str) -> str:
        updated_lines = []
        for line in text.splitlines():
            stripped = line.strip()
            if stripped.startswith("| 最終更新日 |"):
                line = markdown_table_row(["最終更新日", format_date_jp(updated_date)])
            elif stripped.startswith("| TR-"):
                line = replace_markdown_row(line, {8: TM_STATUS})
            updated_lines.append(line)
        return "\n".join(updated_lines) + "\n"

    update_markdown_file(TM_MD_PATH, updater)


def update_tm_workbook(updated_date: date) -> None:
    workbook = load_workbook(TM_XLSX_PATH)

    cover_sheet = workbook["表紙"]
    cover_row = find_label_row(cover_sheet, "最終更新日", 2)
    cover_sheet.cell(cover_row, 3).value = updated_date

    matrix_sheet = workbook["2.対応表"]
    for row_idx in range(4, matrix_sheet.max_row + 1):
        trace_id = matrix_sheet.cell(row_idx, 1).value
        if isinstance(trace_id, str) and trace_id.startswith("TR-"):
            matrix_sheet.cell(row_idx, 9).value = TM_STATUS

    workbook.save(TM_XLSX_PATH)


def expand_evidence_ids(text: str) -> list[str]:
    tokens = [token.strip() for token in text.split(",") if token.strip()]
    expanded: list[str] = []
    range_pattern = re.compile(r"^(.*?)(\d+)\.\.(\d+)$")
    for token in tokens:
        match = range_pattern.match(token)
        if match:
            prefix, start, end = match.groups()
            width = len(start)
            for number in range(int(start), int(end) + 1):
                expanded.append(f"{prefix}{number:0{width}d}")
        else:
            expanded.append(token)
    return expanded


def extract_tc_to_evidence_markdown(path: Path) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if re.match(r"^\|\s*TC-\d{2}\s*\|", stripped):
            cells = split_markdown_row(line)
            mapping[cells[0]] = expand_evidence_ids(cells[5])
    return mapping


def extract_tm_matrix_markdown(path: Path) -> tuple[set[str], set[str], dict[str, list[str]]]:
    requirements: set[str] = set()
    tc_ids: set[str] = set()
    evidence_by_trace: dict[str, list[str]] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if stripped.startswith("| TR-"):
            cells = split_markdown_row(line)
            requirements.add(cells[1])
            for tc_id in re.findall(r"\bTC-\d{2}\b", cells[6]):
                tc_ids.add(tc_id)
            evidence_by_trace[cells[0]] = expand_evidence_ids(cells[7])
    return requirements, tc_ids, evidence_by_trace


def run_tc01(context: TestContext) -> TestResult:
    dm_df = context.actual_dm()
    row = dm_df.loc[dm_df["SUBJID"] == "NE-EN-0001"].iloc[0]
    passed = (
        row["LSVDAT"] != ""
        and row["DTHDAT"] == ""
        and row["RFENDAT"] == row["LSVDAT"]
        and row["DTHFLG"] == ""
    )
    actual = (
        "NE-EN-0001: "
        f"LSVDAT={row['LSVDAT']}, DTHDAT={row['DTHDAT'] or '<空欄>'}, "
        f"RFENDAT={row['RFENDAT']}, DTHFLG={row['DTHFLG'] or '<空欄>'}"
    )
    temp_result = TestResult(
        tc_id="TC-01",
        title="LSVDAT のみ存在する症例",
        requirement_ids=["FR-05", "FR-06"],
        expected="RFENDAT が LSVDAT と一致し、DTHFLG は空欄のままであること。",
        status=PASS if passed else FAIL,
        actual=actual,
        remark="実データの対象症例: NE-EN-0001",
        evidence_ids=["EV-TC-001"],
    )
    temp_result.evidence_docs = {
        "EV-TC-001": build_evidence_markdown(
            temp_result,
            execution_date="",
            mode="runtime",
            extra_sections=[
                ("確認データ行", f"```csv\n{csv_block(pd.DataFrame([row]))}\n```"),
            ],
        )
    }
    return temp_result


def run_tc02(context: TestContext) -> TestResult:
    dm_df = context.actual_dm()
    row = dm_df.loc[dm_df["SUBJID"] == "NE-EN-0015"].iloc[0]
    passed = (
        row["LSVDAT"] == ""
        and row["DTHDAT"] != ""
        and row["RFENDAT"] == row["DTHDAT"]
        and row["DTHFLG"] == "Y"
    )
    actual = (
        "NE-EN-0015: "
        f"LSVDAT={row['LSVDAT'] or '<空欄>'}, DTHDAT={row['DTHDAT']}, "
        f"RFENDAT={row['RFENDAT']}, DTHFLG={row['DTHFLG']}"
    )
    temp_result = TestResult(
        tc_id="TC-02",
        title="DTHDAT のみ存在する症例",
        requirement_ids=["FR-05", "FR-06"],
        expected="RFENDAT が DTHDAT と一致し、DTHFLG が Y になること。",
        status=PASS if passed else FAIL,
        actual=actual,
        remark="実データの対象症例: NE-EN-0015",
        evidence_ids=["EV-TC-002"],
    )
    temp_result.evidence_docs = {
        "EV-TC-002": build_evidence_markdown(
            temp_result,
            execution_date="",
            mode="runtime",
            extra_sections=[
                ("確認データ行", f"```csv\n{csv_block(pd.DataFrame([row]))}\n```"),
            ],
        )
    }
    return temp_result


def run_tc03() -> TestResult:
    synthetic_map = {
        "RGST": pd.DataFrame({"SUBJID": ["TEST-BOTH"], "SEXCD": ["M"], "AGE": ["34"]}),
        "LSVDAT": pd.DataFrame({"SUBJID": ["TEST-BOTH"], "LSVDAT": ["2025-11-10"]}),
        "OC": pd.DataFrame({"SUBJID": ["TEST-BOTH"], "DTHDAT": ["2025-10-31"]}),
    }
    with patched_format_dataset(synthetic_map):
        buffer = io_string()
        with redirect_stdout(buffer):
            dm_df = study_functions.DM()
    warning_text = buffer.getvalue().strip()
    passed = (
        dm_df.iloc[0]["RFENDAT"] == "2025-10-31"
        and dm_df.iloc[0]["DTHFLG"] == "Y"
        and "同时存在 LSVDAT 和 DTHDAT" in warning_text
    )
    actual = (
        "擬似症例 TEST-BOTH: "
        f"RFENDAT={dm_df.iloc[0]['RFENDAT']}, DTHFLG={dm_df.iloc[0]['DTHFLG']}, "
        f"警告出力={'あり' if warning_text else 'なし'}"
    )
    temp_result = TestResult(
        tc_id="TC-03",
        title="LSVDAT と DTHDAT が両方存在する警告ケース",
        requirement_ids=["FR-05", "FR-06", "FR-07"],
        expected="RFENDAT は DTHDAT を優先し、DTHFLG は Y、かつ警告が出力されること。",
        status=PASS if passed else FAIL,
        actual=actual,
        remark="1 行の擬似データセットで確認。",
        evidence_ids=["EV-TC-003"],
    )
    temp_result.evidence_docs = {
        "EV-TC-003": build_evidence_markdown(
            temp_result,
            execution_date="",
            mode="runtime",
            extra_sections=[
                ("確認データ行", f"```csv\n{csv_block(dm_df.iloc[[0]])}\n```"),
                ("取得した警告", f"```text\n{warning_text or '<警告なし>'}\n```"),
            ],
        )
    }
    return temp_result


def run_tc04() -> TestResult:
    synthetic_map = {
        "RGST": pd.DataFrame({"SUBJID": ["TEST-NONE"], "SEXCD": ["F"], "AGE": ["40"]}),
        "LSVDAT": pd.DataFrame({"SUBJID": ["TEST-NONE"], "LSVDAT": [""]}),
        "OC": pd.DataFrame({"SUBJID": ["TEST-NONE"], "DTHDAT": [""]}),
    }
    with patched_format_dataset(synthetic_map):
        buffer = io_string()
        with redirect_stdout(buffer):
            dm_df = study_functions.DM()
    warning_text = buffer.getvalue().strip()
    passed = (
        dm_df.iloc[0]["RFENDAT"] == ""
        and dm_df.iloc[0]["DTHFLG"] == ""
        and "LSVDAT 和 DTHDAT 均为空" in warning_text
    )
    actual = (
        "擬似症例 TEST-NONE: "
        f"RFENDAT={dm_df.iloc[0]['RFENDAT'] or '<空欄>'}, "
        f"DTHFLG={dm_df.iloc[0]['DTHFLG'] or '<空欄>'}, "
        f"警告出力={'あり' if warning_text else 'なし'}"
    )
    temp_result = TestResult(
        tc_id="TC-04",
        title="LSVDAT と DTHDAT が両方空の警告ケース",
        requirement_ids=["FR-05", "FR-06", "FR-07"],
        expected="RFENDAT と DTHFLG が空欄となり、警告が出力されること。",
        status=PASS if passed else FAIL,
        actual=actual,
        remark="1 行の擬似データセットで確認。",
        evidence_ids=["EV-TC-004"],
    )
    temp_result.evidence_docs = {
        "EV-TC-004": build_evidence_markdown(
            temp_result,
            execution_date="",
            mode="runtime",
            extra_sections=[
                ("確認データ行", f"```csv\n{csv_block(dm_df.iloc[[0]])}\n```"),
                ("取得した警告", f"```text\n{warning_text or '<警告なし>'}\n```"),
            ],
        )
    }
    return temp_result


def run_tc05(context: TestContext) -> TestResult:
    actual_filtered = context.actual_tme_registration()
    actual_ok = (
        len(actual_filtered) == 200
        and "EventId" in actual_filtered.columns
        and actual_filtered["EventId"].eq("AT REGISTRATION").all()
    )
    synthetic_source = pd.DataFrame(
        {
            "EventId": ["AT REGISTRATION", "OTHER"],
            "KEEP": ["x", ""],
            "ALL_EMPTY": ["", ""],
            "FLAG": ["Y", ""],
        }
    )
    synthetic_filtered = study_functions.filter_df_by_field(
        synthetic_source, EventId="AT REGISTRATION"
    )
    values_are_strings = (
        synthetic_filtered.apply(lambda series: series.map(lambda value: isinstance(value, str))).all().all()
    )
    synthetic_ok = (
        "ALL_EMPTY" not in synthetic_filtered.columns
        and list(synthetic_filtered.columns) == ["EventId", "KEEP", "FLAG"]
        and values_are_strings
    )
    passed = actual_ok and synthetic_ok
    actual = (
        f"実データ抽出行数={len(actual_filtered)}; "
        f"擬似データ列={list(synthetic_filtered.columns)}; "
        f"擬似データ型={[str(dtype) for dtype in synthetic_filtered.dtypes]}"
    )
    temp_result = TestResult(
        tc_id="TC-05",
        title="filter_df_by_field 正常系",
        requirement_ids=["FR-01", "FR-02", "FR-03"],
        expected="文字列ソースの抽出が成功し、全空列が削除され、出力値がすべて文字列であること。",
        status=PASS if passed else FAIL,
        actual=actual,
        remark="実データの TME と擬似 DataFrame を組み合わせて確認。",
        evidence_ids=["EV-TC-005"],
    )
    temp_result.evidence_docs = {
        "EV-TC-005": build_evidence_markdown(
            temp_result,
            execution_date="",
            mode="runtime",
            extra_sections=[
                ("実データ抽出サンプル", f"```csv\n{csv_block(actual_filtered.head(5))}\n```"),
                ("擬似データ抽出結果", f"```csv\n{csv_block(synthetic_filtered)}\n```"),
            ],
        )
    }
    return temp_result


def run_tc06() -> TestResult:
    captured_error = ""
    passed = False
    try:
        study_functions.filter_df_by_field("TME", NotExists="x")
    except KeyError as exc:
        captured_error = str(exc)
        passed = "NotExists" in captured_error
    except Exception as exc:
        captured_error = f"{type(exc).__name__}: {exc}"
    else:
        captured_error = "例外は発生しませんでした。"

    actual = (
        "存在しないフィールド指定により KeyError が発生。"
        if passed
        else f"期待した KeyError を確認できませんでした。詳細: {captured_error}"
    )
    temp_result = TestResult(
        tc_id="TC-06",
        title="filter_df_by_field 異常系",
        requirement_ids=["FR-01"],
        expected="存在しないフィールド指定時に KeyError が発生すること。",
        status=PASS if passed else FAIL,
        actual=actual,
        remark="実データの TME に存在しないフィールドを指定して確認。",
        evidence_ids=["EV-TC-006"],
    )
    temp_result.evidence_docs = {
        "EV-TC-006": build_evidence_markdown(
            temp_result,
            execution_date="",
            mode="runtime",
            extra_sections=[("取得した例外", f"```text\n{captured_error}\n```")],
        )
    }
    return temp_result


def run_tc07() -> TestResult:
    dm_df = operate_type.DM()
    passed = (
        callable(getattr(operate_type, "DM", None))
        and callable(getattr(operate_type, "filter_df_by_field", None))
        and getattr(operate_type.DM, "__module__", "").endswith("VC_BC05_studyFunctions")
        and getattr(operate_type.filter_df_by_field, "__module__", "").endswith("VC_BC05_studyFunctions")
        and list(dm_df.columns) == ["SUBJID", "SEXCD", "AGE", "LSVDAT", "DTHDAT", "RFENDAT", "DTHFLG"]
    )
    actual = (
        "VC_BC04_operateType から研究固有関数を呼び出し、"
        f"DM 形状={dm_df.shape} を取得。"
    )
    temp_result = TestResult(
        tc_id="TC-07",
        title="VC_BC04_operateType からの上位工程連携",
        requirement_ids=["FR-08"],
        expected="上位工程モジュールから研究固有関数を正常に呼び出せること。",
        status=PASS if passed else FAIL,
        actual=actual,
        remark="関数公開状態と DM の実呼び出しを確認。",
        evidence_ids=["EV-TC-007"],
    )
    temp_result.evidence_docs = {
        "EV-TC-007": build_evidence_markdown(
            temp_result,
            execution_date="",
            mode="runtime",
            extra_sections=[("DM 出力サンプル", f"```csv\n{csv_block(dm_df.head(3))}\n```")],
        )
    }
    return temp_result


def run_tc08() -> TestResult:
    rd_text = RD_MD_PATH.read_text(encoding="utf-8")
    bd_text = BD_MD_PATH.read_text(encoding="utf-8")
    dd_tc_to_evidence = extract_tc_to_evidence_markdown(DD_MD_PATH)
    te_tc_to_evidence = extract_tc_to_evidence_markdown(TE_MD_PATH)
    tm_requirements, tm_tc_ids, tm_evidence_by_trace = extract_tm_matrix_markdown(TM_MD_PATH)

    rd_requirements = set(re.findall(r"\b(?:FR|NFR)-\d{2}\b", rd_text))
    bd_requirements = set(re.findall(r"\b(?:FR|NFR)-\d{2}\b", bd_text))
    rd_functional_requirements = {requirement for requirement in rd_requirements if requirement.startswith("FR-")}
    dd_tc_ids = set(dd_tc_to_evidence)
    te_tc_ids = set(te_tc_to_evidence)
    tm_evidence_flat = {evidence_id for values in tm_evidence_by_trace.values() for evidence_id in values}
    dd_evidence_flat = {evidence_id for values in dd_tc_to_evidence.values() for evidence_id in values}
    te_evidence_flat = {evidence_id for values in te_tc_to_evidence.values() for evidence_id in values}

    issues: list[str] = []
    if rd_requirements != tm_requirements:
        issues.append(f"要件不一致: {sorted(rd_requirements ^ tm_requirements)}")
    if not bd_requirements.issuperset(rd_functional_requirements):
        issues.append(
            f"BD の機能要件カバレッジ不足: {sorted(rd_functional_requirements - bd_requirements)}"
        )
    if dd_tc_ids != te_tc_ids or dd_tc_ids != tm_tc_ids:
        issues.append(
            f"TC 不一致: DD^TE={sorted(dd_tc_ids ^ te_tc_ids)}, DD^TM={sorted(dd_tc_ids ^ tm_tc_ids)}"
        )
    if dd_evidence_flat != te_evidence_flat or not tm_evidence_flat.issuperset(dd_evidence_flat):
        issues.append(
            "証跡ID 不一致: "
            f"DD^TE={sorted(dd_evidence_flat ^ te_evidence_flat)}, "
            f"DD-TM={sorted(dd_evidence_flat - tm_evidence_flat)}"
        )

    passed = not issues
    actual = "トレーサビリティ上の不整合は検出されませんでした。" if passed else "; ".join(issues)
    temp_result = TestResult(
        tc_id="TC-08",
        title="トレーサビリティ照合",
        requirement_ids=["FR-09"],
        expected="RD/BD/DD/TE/TM の要件ID、TC-ID、証跡ID が相互に整合していること。",
        status=PASS if passed else FAIL,
        actual=actual,
        remark="Markdown 文書を機械照合して確認。",
        evidence_ids=["EV-TC-008"],
    )
    temp_result.evidence_docs = {
        "EV-TC-008": build_evidence_markdown(
            temp_result,
            execution_date="",
            mode="runtime",
            extra_sections=[
                (
                    "照合した集合",
                    "```json\n"
                    + json.dumps(
                        {
                            "rd_requirements": sorted(rd_requirements),
                            "bd_requirements": sorted(bd_requirements),
                            "tm_requirements": sorted(tm_requirements),
                            "dd_tc_ids": sorted(dd_tc_ids),
                            "te_tc_ids": sorted(te_tc_ids),
                            "tm_tc_ids": sorted(tm_tc_ids),
                            "dd_evidence_ids": sorted(dd_evidence_flat),
                            "te_evidence_ids": sorted(te_evidence_flat),
                            "tm_evidence_ids": sorted(tm_evidence_flat),
                        },
                        ensure_ascii=False,
                        indent=2,
                    )
                    + "\n```"
                )
            ],
        )
    }
    return temp_result


def run_tc09() -> TestResult:
    rgst_df = pd.DataFrame(
        {
            "SUBJID": [f"TEST-{index:05d}" for index in range(10000)],
            "SEXCD": ["M"] * 10000,
            "AGE": ["34"] * 10000,
        }
    )
    lsvdat_df = pd.DataFrame(
        {
            "SUBJID": [f"TEST-{index:05d}" for index in range(0, 10000, 2)],
            "LSVDAT": ["2025-11-10"] * 5000,
        }
    )
    oc_df = pd.DataFrame(
        {
            "SUBJID": [f"TEST-{index:05d}" for index in range(1, 10000, 2)],
            "DTHDAT": ["2025-10-31"] * 5000,
        }
    )
    synthetic_map = {"RGST": rgst_df, "LSVDAT": lsvdat_df, "OC": oc_df}
    with patched_format_dataset(synthetic_map):
        buffer = io_string()
        start = time.perf_counter()
        with redirect_stdout(buffer):
            dm_df = study_functions.DM()
        elapsed = time.perf_counter() - start
    passed = elapsed <= 3.0 and dm_df.shape == (10000, 7)
    actual = f"10000 行を {elapsed:.6f} 秒で処理し、出力形状は {dm_df.shape} でした。"
    temp_result = TestResult(
        tc_id="TC-09",
        title="1万件擬似データによる性能試験",
        requirement_ids=["NFR-01"],
        expected="1万件の擬似データで DM 生成が 3 秒以内に完了すること。",
        status=PASS if passed else FAIL,
        actual=actual,
        remark="擬似データのみ使用し、セットアップ時間は測定対象外。",
        evidence_ids=["EV-PERF-001"],
    )
    temp_result.evidence_docs = {
        "EV-PERF-001": build_evidence_markdown(
            temp_result,
            execution_date="",
            mode="runtime",
            extra_sections=[
                (
                    "擬似データ件数",
                    "```json\n"
                    + json.dumps(
                        {"RGST": len(rgst_df), "LSVDAT": len(lsvdat_df), "OC": len(oc_df)},
                        ensure_ascii=False,
                        indent=2,
                    )
                    + "\n```"
                )
            ],
        )
    }
    return temp_result


def run_tc10_pre(context: TestContext) -> TestResult:
    dm_df = context.actual_dm()
    unexpected_columns = sorted(set(dm_df.columns) - ALLOWED_DM_COLUMNS)
    te_text = TE_MD_PATH.read_text(encoding="utf-8")
    tm_text = TM_MD_PATH.read_text(encoding="utf-8")
    passed = (
        not unexpected_columns
        and "## 6. 変更履歴" in te_text
        and "## 5. 変更履歴" in tm_text
    )
    actual = (
        f"DM 列={list(dm_df.columns)}; "
        f"想定外列={unexpected_columns or 'なし'}; "
        "TE/TM に変更履歴セクションが存在。"
    )
    return TestResult(
        tc_id="TC-10",
        title="セキュリティおよび監査性確認",
        requirement_ids=["NFR-04", "NFR-05"],
        expected="DM 出力に不要な個人情報列が含まれず、TE/TM に変更履歴と保存済み証跡リンクが保持されること。",
        status=PASS if passed else FAIL,
        actual=actual,
        remark="文書更新後に証跡リンクの存在確認を追加で行う。",
        evidence_ids=["EV-SEC-001", "EV-AUD-001"],
    )


def finalize_tc10_post_save(result: TestResult) -> None:
    te_text = TE_MD_PATH.read_text(encoding="utf-8")
    tm_text = TM_MD_PATH.read_text(encoding="utf-8")
    evidence_link_checks = []
    for evidence_id, evidence_path in result.evidence_paths.items():
        evidence_link_checks.append(
            {
                "evidence_id": evidence_id,
                "in_te": evidence_path in te_text,
                "in_tm": evidence_id in tm_text,
            }
        )
    links_present = all(check["in_te"] and check["in_tm"] for check in evidence_link_checks)
    passed = result.status == PASS and links_present
    result.status = PASS if passed else FAIL
    result.actual = (
        result.actual
        + " 保存後証跡リンク確認: "
        + ", ".join(
            f"{check['evidence_id']}[TE={check['in_te']},TM={check['in_tm']}]"
            for check in evidence_link_checks
        )
    )
    result.remark = "DM 出力は保存前に確認し、TE/TM の証跡リンクは保存後に確認。"
    result.evidence_docs = {
        "EV-SEC-001": build_evidence_markdown(
            TestResult(
                tc_id=result.tc_id,
                title="セキュリティ観点の出力確認",
                requirement_ids=["NFR-04"],
                expected="DM 出力に不要な個人情報列が含まれないこと。",
                status=result.status,
                actual=result.actual,
                remark=result.remark,
                evidence_ids=["EV-SEC-001"],
            ),
            execution_date="",
            mode="runtime",
            extra_sections=[("確認範囲", "許容された DM 列のみを判定対象とした。")],
        ),
        "EV-AUD-001": build_evidence_markdown(
            TestResult(
                tc_id=result.tc_id,
                title="監査観点の証跡リンク確認",
                requirement_ids=["NFR-05"],
                expected="TE/TM に変更履歴セクションと保存済み証跡リンクが保持されること。",
                status=result.status,
                actual=result.actual,
                remark=result.remark,
                evidence_ids=["EV-AUD-001"],
            ),
            execution_date="",
            mode="runtime",
            extra_sections=[
                (
                    "証跡リンク確認結果",
                    "```json\n" + json.dumps(evidence_link_checks, ensure_ascii=False, indent=2) + "\n```",
                )
            ],
        ),
    }


def assign_evidence_paths(results: list[TestResult], evidence_dir: Path | None) -> None:
    if evidence_dir is None:
        return
    for result in results:
        for evidence_id in result.evidence_ids:
            evidence_path = evidence_dir / f"{evidence_id}.md"
            result.evidence_paths[evidence_id] = evidence_relpath(evidence_path)


def stamp_evidence_docs(results: list[TestResult], execution_date: str, mode: str) -> None:
    mode_display = {"save": "保存実行", "dry-run": "ドライラン"}.get(mode, mode)
    for result in results:
        for evidence_id in list(result.evidence_docs):
            result.evidence_docs[evidence_id] = result.evidence_docs[evidence_id].replace(
                "- 実施日: ", f"- 実施日: {execution_date}"
            ).replace("- 実行モード: runtime", f"- 実行モード: {mode_display}")


def run_all_tests() -> tuple[list[TestResult], TestContext]:
    context = TestContext()
    results = [
        run_tc01(context),
        run_tc02(context),
        run_tc03(),
        run_tc04(),
        run_tc05(context),
        run_tc06(),
        run_tc07(),
        run_tc08(),
        run_tc09(),
        run_tc10_pre(context),
    ]
    return results, context


def results_by_tc(results: list[TestResult]) -> dict[str, TestResult]:
    return {result.tc_id: result for result in results}


def defect_row_for_results(results: list[TestResult], execution_date: date) -> list[str]:
    failures = [result for result in results if result.status == FAIL]
    if failures:
        failure_ids = ", ".join(result.tc_id for result in failures)
        return [
            "DEF-001",
            format_date_iso(execution_date),
            f"Fail recorded in {failure_ids}",
            "VC_BC05_studyFunctions / CombineProcess evidence",
            "Record only; no production code fix performed in this run.",
            "Open",
        ]
    return ["DEF-001", format_date_iso(execution_date), "なし", "-", "-", "Closed"]


def summary_text_for_results(results: list[TestResult], evidence_dir: Path | None) -> str:
    passed = sum(result.status == PASS for result in results)
    failed = sum(result.status == FAIL for result in results)
    base = f"{len(results)}件中 {passed} Pass / {failed} Fail"
    if evidence_dir is not None:
        return (
            f"{base} "
            f"(証跡: {evidence_relpath(evidence_dir / 'summary_ja.md')}, "
            f"JSON: {evidence_relpath(evidence_dir / 'summary.json')})"
        )
    return base


def write_evidence(results: list[TestResult], evidence_dir: Path) -> None:
    evidence_dir.mkdir(parents=True, exist_ok=True)
    for result in results:
        for evidence_id in result.evidence_ids:
            content = result.evidence_docs[evidence_id]
            (evidence_dir / f"{evidence_id}.md").write_text(content, encoding="utf-8")


def build_summary(
    results: list[TestResult],
    execution_dt: datetime,
    mode: str,
    environment: str,
    evidence_dir: Path | None,
    post_save_validation: dict[str, Any] | None = None,
) -> RunSummary:
    pass_count = sum(result.status == PASS for result in results)
    fail_count = sum(result.status == FAIL for result in results)
    return RunSummary(
        generated_at=execution_dt.isoformat(),
        mode=mode,
        execution_date=format_date_iso(execution_dt.date()),
        environment=environment,
        evidence_directory=evidence_relpath(evidence_dir) if evidence_dir else None,
        pass_count=pass_count,
        fail_count=fail_count,
        overall_status=PASS if fail_count == 0 else FAIL,
        results=[
            {
                "tc_id": result.tc_id,
                "title": result.title,
                "requirements": result.requirement_ids,
                "expected": result.expected,
                "status": result.status,
                "actual": result.actual,
                "remark": result.remark,
                "evidence_ids": result.evidence_ids,
                "evidence_paths": result.evidence_paths,
            }
            for result in results
        ],
        post_save_validation=post_save_validation or {},
    )


def save_summary(summary: RunSummary, evidence_dir: Path) -> None:
    (evidence_dir / "summary.json").write_text(
        json.dumps(asdict(summary), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def save_summary_markdown(summary: RunSummary, evidence_dir: Path) -> None:
    lines = [
        "# 実施結果サマリー",
        "",
        f"- 実施日時: {summary.generated_at}",
        f"- 実行モード: {'保存実行' if summary.mode == 'save' else 'ドライラン'}",
        f"- 実施日: {summary.execution_date}",
        f"- 実施環境: {summary.environment}",
        f"- 証跡フォルダ: {summary.evidence_directory or '未保存'}",
        f"- 合計件数: {summary.pass_count + summary.fail_count}",
        f"- Pass: {summary.pass_count}",
        f"- Fail: {summary.fail_count}",
        f"- 総合判定: {summary.overall_status}",
        "",
        "## ケース別結果",
        "",
    ]
    for result in summary.results:
        lines.extend(
            [
                f"### {result['tc_id']} {result['title']}",
                f"- 判定: {result['status']}",
                f"- 実測結果: {result['actual']}",
                f"- 備考: {result['remark'] or 'なし'}",
                "",
            ]
        )
    if summary.post_save_validation:
        lines.extend(
            [
                "## 保存後検証",
                "",
                f"- 判定: {'Pass' if summary.post_save_validation.get('passed') else 'Fail'}",
                f"- 指摘: {summary.post_save_validation.get('issues') or 'なし'}",
                "",
            ]
        )
    (evidence_dir / "summary_ja.md").write_text("\n".join(lines), encoding="utf-8")


def validate_saved_state(results: list[TestResult]) -> dict[str, Any]:
    issues: list[str] = []
    result_map = results_by_tc(results)

    dd_text = DD_MD_PATH.read_text(encoding="utf-8")
    te_text = TE_MD_PATH.read_text(encoding="utf-8")
    tm_text = TM_MD_PATH.read_text(encoding="utf-8")

    for tc_id, result in result_map.items():
        dd_pattern = re.compile(
            rf"^\|\s*{re.escape(tc_id)}\s*\|.*\|\s*{re.escape(result.status_for_docs())}\s*\|$",
            re.MULTILINE,
        )
        if not dd_pattern.search(dd_text):
            issues.append(f"DD markdown does not contain status for {tc_id}.")
        if sanitize_cell_text(result.actual) not in te_text:
            issues.append(f"TE markdown does not contain actual result for {tc_id}.")
        if result.evidence_cell_markdown() and result.evidence_cell_markdown() not in te_text:
            issues.append(f"TE markdown does not contain evidence path for {tc_id}.")

    if TM_STATUS not in tm_text:
        issues.append("TM markdown was not updated to In Progress.")

    dd_wb = load_workbook(DD_XLSX_PATH, data_only=False)
    dd_sheet = dd_wb["4.テスト仕様"]
    for tc_id, result in result_map.items():
        row_idx = find_row_by_first_value(dd_sheet, tc_id, 1)
        if normalize_doc_text(str(dd_sheet.cell(row_idx, 7).value or "")) != normalize_doc_text(result.status_for_docs()):
            issues.append(f"DD workbook mismatch for {tc_id}.")

    te_wb = load_workbook(TE_XLSX_PATH, data_only=False)
    te_sheet = te_wb["2.テスト結果詳細"]
    for tc_id, result in result_map.items():
        row_idx = find_row_by_first_value(te_sheet, tc_id, 1)
        actual_value = normalize_doc_text(str(te_sheet.cell(row_idx, 4).value or ""))
        evidence_value = normalize_doc_text(str(te_sheet.cell(row_idx, 7).value or ""))
        if actual_value != normalize_doc_text(result.actual):
            issues.append(f"TE workbook actual mismatch for {tc_id}.")
        if evidence_value != normalize_doc_text(result.evidence_cell_workbook()):
            issues.append(f"TE workbook evidence mismatch for {tc_id}.")

    tm_wb = load_workbook(TM_XLSX_PATH, data_only=False)
    tm_sheet = tm_wb["2.対応表"]
    for row_idx in range(4, tm_sheet.max_row + 1):
        trace_id = tm_sheet.cell(row_idx, 1).value
        if isinstance(trace_id, str) and trace_id.startswith("TR-"):
            status_value = normalize_doc_text(str(tm_sheet.cell(row_idx, 9).value or ""))
            if status_value != normalize_doc_text(TM_STATUS):
                issues.append(f"TM workbook status mismatch for {trace_id}.")

    return {"passed": not issues, "issues": issues}


def persist_documents_final(
    results: list[TestResult],
    execution_dt: datetime,
    environment: str,
    evidence_dir: Path | None,
) -> dict[str, Any]:
    execution_date = execution_dt.date()
    result_map = results_by_tc(results)
    summary_text = summary_text_for_results(results, evidence_dir)
    defect_row = defect_row_for_results(results, execution_date)

    update_dd_markdown(result_map, execution_date)
    update_dd_workbook(result_map, execution_date)
    update_tm_markdown(execution_date)
    update_tm_workbook(execution_date)
    update_te_markdown(result_map, execution_date, environment, summary_text, defect_row)
    update_te_workbook(result_map, execution_date, environment, summary_text, defect_row)

    return {"summary_text": summary_text, "defect_row": defect_row}


def create_evidence_dir(execution_dt: datetime) -> Path:
    folder_name = f"evidence-{execution_dt.strftime('%Y%m%d-%H%M%S')}"
    return EVIDENCE_ROOT / folder_name


def main() -> int:
    args = parse_args()
    mode = "save" if args.save else "dry-run"
    execution_dt = now_local()
    environment = environment_string()

    results, _ = run_all_tests()
    evidence_dir = create_evidence_dir(execution_dt) if args.save else None
    assign_evidence_paths(results, evidence_dir)

    if args.save and evidence_dir is not None:
        persist_documents_final(results, execution_dt, environment, evidence_dir)
        finalize_tc10_post_save(results_by_tc(results)["TC-10"])
        persist_documents_final(results, execution_dt, environment, evidence_dir)
        stamp_evidence_docs(results, format_date_iso(execution_dt.date()), mode)
        write_evidence(results, evidence_dir)
        validation = validate_saved_state(results)
        summary = build_summary(results, execution_dt, mode, environment, evidence_dir, validation)
        save_summary(summary, evidence_dir)
        save_summary_markdown(summary, evidence_dir)
    else:
        stamp_evidence_docs(results, format_date_iso(execution_dt.date()), mode)
        summary = build_summary(results, execution_dt, mode, environment, evidence_dir)

    print(json.dumps(asdict(summary), ensure_ascii=False, indent=2))
    return 0 if summary.overall_status == PASS else 1


if __name__ == "__main__":
    raise SystemExit(main())
