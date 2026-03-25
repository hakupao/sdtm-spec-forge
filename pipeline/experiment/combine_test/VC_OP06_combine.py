"""
VAPORCONE 项目数据转换模块

该模块负责读取Excel文件中的Combine_process工作表，并根据Step名执行相应的转换函数：
- 读取Combine_process工作表配置
- 根据Step名调用不同的转换函数
- 处理各种参数类型
- 输出转换后的CSV文件到COMBINE目录

支持的操作类型：
- MERGE: 合并多个文件
- SORT: 数据排序
- CONCAT: 上下连接多个文件
"""

import sys
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

CURRENT_DIR = Path(__file__).resolve().parent


def find_repo_root(start: Path) -> Path:
    for candidate in (start, *start.parents):
        if (candidate / 'VC_BC01_constant.py').exists():
            return candidate
    return start


REPO_ROOT = find_repo_root(CURRENT_DIR)
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from VC_BC02_baseUtils import find_latest_timestamped_path, get_cell_value
from VC_BC03_fetchConfig import getSheetSetting
from VC_BC01_constant import (
    COL_FILENAME,
    COL_STARTINGROW,
    COL_STEP,
    COMBINE_PROCESS_SHEET_NAME,
    CONFIG_NAME,
    EXTENSION,
    FOLDER_FORMAT,
    FORMAT_PATH,
    PREFIX_F,
    SPECIFIC_PATH,
    STUDY_ID,
)

# 输出路径配置
FOLDER_COMBINE = 'TEST_COMBINE'
EXPERIMENT_ROOT = CURRENT_DIR
COMBINE_PATH = EXPERIMENT_ROOT / FOLDER_COMBINE
COMBINE_TRANSFER_FILE_PATH = COMBINE_PATH / 'combine_dataset'
EXPERIMENT_FORMAT_PATH = EXPERIMENT_ROOT / FOLDER_FORMAT

def resolve_existing_path(*candidates):
    "Return the first existing path from candidates, or None."
    for candidate in candidates:
        path = Path(candidate)
        if path.exists():
            return path
    return None

# 支持的Step类型
SUPPORTED_STEPS = {
    'MERGE': 'merge_files',
    'SORT': 'sort_data',
    'CONCAT': 'concat_files'
}

def get_combine_process_info(workbook, sheetSetting):
    """
    从Combine_process工作表中读取转换配置信息
    
    参数:
    - workbook: Excel工作簿对象
    - sheetSetting: 工作表设置字典
    
    返回:
    - list: 转换任务列表，每个任务包含文件名、Step名和参数
    """
    if COMBINE_PROCESS_SHEET_NAME not in sheetSetting:
        print(f"警告: 未找到 {COMBINE_PROCESS_SHEET_NAME} 工作表配置")
        return []
    
    combine_sheet = workbook[COMBINE_PROCESS_SHEET_NAME]
    combine_sheet_setting = sheetSetting[COMBINE_PROCESS_SHEET_NAME]
    
    # 获取列索引 - 根据SheetSetting配置
    col_filename = combine_sheet_setting.get(COL_FILENAME, 0)  # 从配置读取FILENAME列索引
    col_step = combine_sheet_setting.get(COL_STEP, 1)          # 从配置读取STEP列索引
    col_parameters = combine_sheet_setting.get('PARA1', 2)     # 从配置读取PARA1列索引，如果没有则默认为2
    
    # 动态确定最大列数
    max_col = 0
    for cell in combine_sheet[1]:  # 第一行
        if cell.value is None:
            break
        max_col += 1
    
    print(f"调试信息: 最大列数 = {max_col}")
    
    tasks = []
    for row in combine_sheet.iter_rows(
        min_row=combine_sheet_setting[COL_STARTINGROW], 
        min_col=1, 
        max_col=max_col, 
        values_only=True
    ):
        if not any(row):
            break
        
        filename = get_cell_value(row, col_filename)
        step = get_cell_value(row, col_step)
        parameters = {}
        
        print(f"调试信息: 读取行 - FILENAME: {filename}, STEP: {step}")
        
        # 解析C列之后的参数，直接使用冒号后面的值
        for col_idx in range(col_parameters, max_col):
            if col_idx < len(row) and row[col_idx] is not None:
                param_value = row[col_idx]
                if param_value:
                    # 参数值格式如："连接文件：SS_A, SS_A_UNS"
                    # 直接提取冒号后面的内容作为参数值
                    clean_param_value = extract_parameter_value(param_value)
                    if clean_param_value:
                        # 使用列索引作为参数名（PARA1, PARA2, PARA3...）
                        param_name = f"PARA{col_idx - col_parameters + 1}"
                        parameters[param_name] = clean_param_value
                        print(f"调试信息: 参数 {param_name} = {clean_param_value}")
        
        if filename and step:
            # 清理文件名，去掉"输出："前缀
            clean_filename = extract_parameter_value(filename)
            # 清理步骤名，去掉汉字前缀
            clean_step = extract_step_name(step)
            
            print(f"调试信息: 清理后 - FILENAME: {clean_filename}, STEP: {clean_step}")
            
            tasks.append({
                'filename': clean_filename,
                'step': clean_step,
                'parameters': parameters
            })
    
    print(f"调试信息: 总共找到 {len(tasks)} 个任务")
    return tasks


def extract_parameter_value(param_value):
    """
    从参数值中提取干净的值（去掉汉字前缀）
    
    参数:
    - param_value: 原始参数值（如："连接文件：SS_A, SS_A_UNS"）
    
    返回:
    - str: 清理后的值（如："SS_A, SS_A_UNS"）
    """
    if not param_value:
        return ""
    
    # 查找冒号位置
    colon_pos = param_value.find('：')
    if colon_pos == -1:
        colon_pos = param_value.find(':')
    
    if colon_pos != -1:
        # 返回冒号后面的内容
        return param_value[colon_pos + 1:].strip()
    else:
        # 如果没有冒号，直接返回原值
        return param_value.strip()

def extract_step_name(step_name):
    """
    从步骤名中提取干净的步骤类型（去掉汉字前缀）
    
    参数:
    - step_name: 原始步骤名（如："上下连接：CONCAT"）
    
    返回:
    - str: 清理后的步骤类型（如："CONCAT"）
    """
    if not step_name:
        return ""
    
    # 查找冒号位置
    colon_pos = step_name.find('：')
    if colon_pos == -1:
        colon_pos = step_name.find(':')
    
    if colon_pos != -1:
        # 返回冒号后面的内容
        return step_name[colon_pos + 1:].strip()
    else:
        # 如果没有冒号，直接返回原值
        return step_name.strip()

def concat_files(parameters, input_data):
    """
    上下连接多个文件（垂直合并）
    
    参数:
    - parameters: 包含连接配置的参数字典
    - input_data: 输入数据字典
    
    返回:
    - DataFrame: 连接后的数据
    """
    try:
        # 获取要连接的文件列表 - 从PARA1获取
        source_files = parameters.get('PARA1', '').split(',')
        source_files = [f.strip() for f in source_files if f.strip()]
        
        if not source_files:
            print("错误: 未指定连接文件")
            return pd.DataFrame()
        
        dataframes = []
        for file in source_files:
            if file in input_data:
                dataframes.append(input_data[file])
                print(f"  ✓ 加载文件: {file} ({len(input_data[file])} 行)")
            else:
                print(f"  ✗ 未找到文件: {file}")
        
        if not dataframes:
            print("错误: 没有可连接的数据")
            return pd.DataFrame()
        
        # 执行垂直连接（上下连接）
        result = pd.concat(dataframes, ignore_index=True)
        
        print(f"成功连接 {len(dataframes)} 个文件，结果行数: {len(result)}")
        return result
        
    except Exception as e:
        print(f"连接文件时出错: {e}")
        return pd.DataFrame()

def sort_data(parameters, input_data):
    """
    对数据进行排序
    
    参数:
    - parameters: 包含排序配置的参数字典
    - input_data: 输入数据字典
    
    返回:
    - DataFrame: 排序后的数据
    """
    try:
        source_file = parameters.get('PARA1', '')  # 排序文件
        if not source_file or source_file not in input_data:
            print(f"错误: 未找到排序文件 {source_file}")
            return pd.DataFrame()
        
        df = input_data[source_file].copy()
        sort_fields = parameters.get('PARA2', '').split(',')  # 排序字段
        sort_fields = [f.strip() for f in sort_fields if f.strip()]
        
        if not sort_fields:
            print("错误: 未指定排序字段")
            return df
        
        # 检查排序字段是否存在
        available_sort_fields = [f for f in sort_fields if f in df.columns]
        if not available_sort_fields:
            print("错误: 指定的排序字段都不存在")
            return df
        
        # 获取排序方式
        sort_order = parameters.get('PARA3', 'Y,Y').split(',')  # 排序方式
        sort_order = [order.strip().upper() == 'Y' for order in sort_order]
        
        # 如果排序方式数量不匹配，使用默认值
        while len(sort_order) < len(available_sort_fields):
            sort_order.append(True)  # 默认升序
        
        # 截取到可用字段数量
        sort_order = sort_order[:len(available_sort_fields)]
        
        # 获取空值位置处理
        null_position = parameters.get('PARA4', 'first').lower()  # 空值位置
        
        # 执行排序
        df = df.sort_values(by=available_sort_fields, ascending=sort_order, na_position=null_position)
        
        print(f"数据排序完成: {available_sort_fields}")
        print(f"排序方式: {sort_order}")
        print(f"空值位置: {null_position}")
        return df
        
    except Exception as e:
        print(f"排序数据时出错: {e}")
        return pd.DataFrame()

def merge_files(parameters, input_data):
    """
    结合两个文件（水平合并）
    
    参数:
    - parameters: 包含结合配置的参数字典
    - input_data: 输入数据字典
    
    返回:
    - DataFrame: 结合后的数据
    """
    try:
        # 获取要结合的文件列表
        source_files = parameters.get('PARA1', '').split(',')  # 结合文件
        source_files = [f.strip() for f in source_files if f.strip()]
        
        if len(source_files) < 2:
            print("错误: 结合操作需要至少2个文件")
            return pd.DataFrame()
        
        # 获取结合方式
        merge_method = parameters.get('PARA2', 'outer').lower()  # 结合方式
        
        # 获取连接键
        join_keys = parameters.get('PARA3', '').split(',')  # 连接键
        join_keys = [f.strip() for f in join_keys if f.strip()]
        
        if not join_keys:
            print("错误: 未指定连接键")
            return pd.DataFrame()
        
        # 检查文件是否存在
        if source_files[0] not in input_data or source_files[1] not in input_data:
            print(f"错误: 未找到指定的文件")
            return pd.DataFrame()
        
        left_df = input_data[source_files[0]]
        right_df = input_data[source_files[1]]
        
        print(f"  ✓ 左表: {source_files[0]} ({len(left_df)} 行)")
        print(f"  ✓ 右表: {source_files[1]} ({len(right_df)} 行)")
        print(f"  ✓ 连接键: {join_keys}")
        print(f"  ✓ 结合方式: {merge_method}")
        
        # 执行结合
        result = pd.merge(left_df, right_df, on=join_keys, how=merge_method)
        
        print(f"结合完成，结果行数: {len(result)}")
        return result
        
    except Exception as e:
        print(f"结合文件时出错: {e}")
        return pd.DataFrame()

def execute_step(step_name, parameters, input_data):
    """
    根据Step名执行相应的转换函数
    
    参数:
    - step_name: Step名称
    - parameters: 参数字典
    - input_data: 输入数据字典
    
    返回:
    - DataFrame: 转换后的数据
    """
    step_name_upper = step_name.upper()
    
    if step_name_upper in SUPPORTED_STEPS:
        function_name = SUPPORTED_STEPS[step_name_upper]
        function = globals().get(function_name)
        
        if function:
            print(f"执行 {step_name} 步骤...")
            return function(parameters, input_data)
        else:
            print(f"错误: 未找到函数 {function_name}")
            return pd.DataFrame()
    else:
        print(f"错误: 不支持的Step类型 {step_name}")
        print(f"支持的Step类型: {', '.join(SUPPORTED_STEPS.keys())}")
        return pd.DataFrame()

def main():
    """
    主函数，执行数据转换流程
    """
    print(f"开始数据转换处理...")
    start_time = time.time()
    
    # 调试路径常量
    print(f"调试信息: 原始 SPECIFIC_PATH = {SPECIFIC_PATH}")
    print(f"调试信息: 实验目录 = {EXPERIMENT_ROOT}")
    print(f"调试信息: 输出根目录 = {COMBINE_PATH}")
    print(f"调试信息: 输出传输目录 = {COMBINE_TRANSFER_FILE_PATH}")
        
    # 创建输出目录
    print(f"尝试创建输出目录: {COMBINE_TRANSFER_FILE_PATH}")
    
    try:
        COMBINE_TRANSFER_FILE_PATH.mkdir(parents=True, exist_ok=True)
        
        # 生成带时间戳的目录名
        timestamp = time.strftime("%Y%m%d%H%M%S")
        actual_combine_path = COMBINE_TRANSFER_FILE_PATH / f'combine_dataset-{timestamp}'
        
        # 创建最终的输出目录
        actual_combine_path.mkdir(parents=True, exist_ok=True)
        
        print(f'✓ 成功创建输出目录: {actual_combine_path}')
        print(f'输出目录绝对路径: {actual_combine_path.resolve()}')
        
    except Exception as e:
        print(f"✗ 创建输出目录失败: {e}")
        print(f"错误类型: {type(e).__name__}")
        print(f"尝试创建路径: {COMBINE_TRANSFER_FILE_PATH}")
        return
    
    config_path = resolve_existing_path(EXPERIMENT_ROOT / CONFIG_NAME, Path(SPECIFIC_PATH) / CONFIG_NAME)
    if not config_path:
        print(f"错误: 未找到配置文件 {CONFIG_NAME}")
        return
    
    print(f"使用配置文件: {config_path}")
    workbook = load_workbook(filename=str(config_path))
    sheetSetting = getSheetSetting(workbook)
    
    # 读取Combine_process工作表配置
    tasks = get_combine_process_info(workbook, sheetSetting)
    
    if not tasks:
        print("未找到转换任务，程序退出")
        return
    
    print(f"找到 {len(tasks)} 个转换任务")
    
    # 获取格式化数据作为输入
    format_base_path = resolve_existing_path(EXPERIMENT_FORMAT_PATH, Path(FORMAT_PATH))
    if not format_base_path:
        print("错误: 未找到格式化数据根目录")
        return
    
    actual_format_path_str = find_latest_timestamped_path(str(format_base_path), 'format_dataset')
    actual_format_path = Path(actual_format_path_str)
    if not actual_format_path.exists():
        print(f"错误: 未找到格式化数据目录: {actual_format_path}")
        return
    
    print(f'使用格式化数据路径: {actual_format_path}')
    
    # 读取所有格式化数据文件
    input_data = {}
    for file_path in actual_format_path.iterdir():
        if file_path.is_file():
            file_name = file_path.name
            if file_name.startswith(PREFIX_F) and file_name.endswith(EXTENSION):
                short_file_name = file_name.removeprefix(PREFIX_F).removesuffix(EXTENSION)
                try:
                    input_data[short_file_name] = pd.read_csv(file_path, dtype=str, na_filter=False)
                    print(f"加载文件: {short_file_name} ({len(input_data[short_file_name])} 行)")
                except Exception as e:
                    print(f"警告: 无法加载文件 {file_name}: {e}")
    
    # 执行Combine_process工作表中定义的转换任务
    print("\n=== 执行Combine_process工作表转换任务 ===")
    for i, task in enumerate(tasks, 1):
        print(f"\n=== 执行任务 {i}/{len(tasks)} ===")
        print(f"文件名: {task['filename']}")
        print(f"Step: {task['step']}")
        print(f"参数: {task['parameters']}")
        
        task_start = time.time()
        
        # 执行转换
        result_df = execute_step(task['step'], task['parameters'], input_data)
        
        if not result_df.empty:
            # 保存结果 - 添加F-前缀
            output_file = f'{PREFIX_F}{task["filename"]}{EXTENSION}'
            output_path = actual_combine_path / output_file
            
            print(f"调试信息: 准备保存文件")
            print(f"调试信息: 输出文件名: {output_file}")
            print(f"调试信息: 完整输出路径: {output_path}")
            print(f"调试信息: 输出目录是否存在: {actual_combine_path.exists()}")
            
            try:
                result_df.to_csv(output_path, index=False, encoding='utf-8-sig')
                print(f"✓ 转换完成: {output_file} ({len(result_df)} 行)")
                print(f"✓ 文件已保存到: {output_path}")
                
                # 验证文件是否真的保存了
                if output_path.exists():
                    file_size = output_path.stat().st_size
                    print(f"✓ 文件保存验证成功，大小: {file_size} 字节")
                else:
                    print(f"✗ 文件保存验证失败，文件不存在: {output_path}")
                
                # 更新输入数据，供后续任务使用
                input_data[task['filename']] = result_df
                
            except Exception as e:
                print(f"✗ 保存文件失败: {e}")
                print(f"✗ 错误详情: {type(e).__name__}: {str(e)}")
        else:
            print(f"✗ 转换失败或结果为空")
        
        task_time = time.time() - task_start
        print(f"任务耗时: {task_time:.2f} 秒")
    
    total_time = time.time() - start_time
    print(f"\n=== 转换完成 ===")
    print(f"总耗时: {total_time:.2f} 秒")
    print(f"输出目录: {actual_combine_path}")

if __name__ == "__main__":
    print(f'Study:{STUDY_ID} 数据转换处理开始...')
    main()
    print(f'Study:{STUDY_ID} 数据转换处理完成。')
